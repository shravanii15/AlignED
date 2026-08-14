"""
services/reports_excel.py -- real, formatted .xlsx export for the
Program Explorer page.

A CSV is just raw text with commas; it fundamentally can't look
"professional" no matter how the columns are chosen. This builds a real
Excel file -- colored header, bold title block, color-coded priority
cells, sensible column widths, a frozen header row -- while staying just
as easy to sort/filter/re-use as a CSV would be.
"""

import datetime
import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from utils.constants import TIER_FILL_HEX


def build_excel_report(university, program_name, course_count, recs_df):
    wb = Workbook()
    ws = wb.active
    ws.title = "Recommendations"

    navy = "1E3A8A"
    ws.merge_cells("A1:G1")
    ws["A1"] = "AlignED -- Curriculum Gap Report"
    ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=navy)
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:G2")
    ws["A2"] = f"{university} -- {program_name}"
    ws["A2"].font = Font(size=11, italic=True, color="475569")

    ws.merge_cells("A3:G3")
    ws["A3"] = f"{course_count} courses analyzed  |  {len(recs_df)} significant gaps  |  Generated {datetime.date.today().isoformat()}"
    ws["A3"].font = Font(size=9, color="94A3B8")

    headers = ["Rank", "Skill", "Priority", "Program Coverage", "Market Demand", "Gap (points)", "Demand Trend"]
    header_row = 5
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.alignment = Alignment(horizontal="center")

    for i, (_, row) in enumerate(recs_df.iterrows(), start=1):
        r = header_row + i
        ws.cell(row=r, column=1, value=i).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=2, value=row["canonical_name"])
        tier_cell = ws.cell(row=r, column=3, value=row["priority_tier"].title())
        tier_cell.fill = PatternFill("solid", fgColor=TIER_FILL_HEX.get(row["priority_tier"], "E2E8F0"))
        tier_cell.alignment = Alignment(horizontal="center")
        tier_cell.font = Font(bold=True)
        cov_cell = ws.cell(row=r, column=4, value=row["program_coverage_rate"])
        cov_cell.number_format = "0.0%"
        dem_cell = ws.cell(row=r, column=5, value=row["market_demand_rate"])
        dem_cell.number_format = "0.0%"
        gap_cell = ws.cell(row=r, column=6, value=row["gap_value"])
        gap_cell.number_format = "+0.0%;-0.0%"
        ws.cell(row=r, column=7, value=row["trend_label"].replace("no clear trend", "Flat").title())

    widths = [7, 22, 12, 18, 16, 14, 14]
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = f"A{header_row + 1}"

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
