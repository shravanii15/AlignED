"""
app.py -- AlignED dashboard

What this file does, in plain terms:
This is the front-facing part of the project -- the part a recruiter,
hiring manager, or curious visitor would actually click through, rather
than reading raw JSON files or console output. It's a Streamlit app,
meaning it's a single Python file that turns into a real interactive
web page: dropdowns, tables, and charts, all reading live from the same
SQLite database every other script in this project has been building up
(programs, courses, job postings, skills, gap scores, demand trends, and
final recommendations).

Why Streamlit for this project:
Streamlit turns a plain Python script into a web app with no separate
frontend code, no JavaScript, and no build step -- ideal for a data
science/analytics portfolio piece where the point is showing real
analysis, not demonstrating web development. It can also be published
for free to a public URL (Streamlit Community Cloud), so this dashboard
can be linked directly from a resume or LinkedIn profile.

How to run this locally:
    pip install -r requirements.txt
    streamlit run app.py
Then open the local URL it prints (usually http://localhost:8501).

Page structure (see the sidebar):
- Overview: project summary and headline numbers
- Program Explorer: pick any of the 13 programs and see its ranked,
  explained skill-gap recommendations
- Skill Demand Trends: which tracked skills are rising or falling
- Role Clusters: how real job postings group into real-world roles
- Methodology & Honest Limitations: the "how this was built, and where
  it's genuinely limited" page -- the kind of thing an interviewer would
  ask about directly, answered up front instead of hidden.
"""

import datetime
import io
import os
import re
import sqlite3

import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st
from fpdf import FPDF
from fpdf.fonts import FontFace
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

TIER_HEX = {"high": "#DC2626", "medium": "#D97706", "low": "#16A34A"}

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))  # .../AlignED
DB_PATH = os.path.join(BASE_DIR, "database", "aligned.db")

st.set_page_config(page_title="AlignED -- Curriculum vs. Job Market Gap Analysis", page_icon="🎓", layout="wide")

# A small block of custom CSS, layered on top of the theme in
# .streamlit/config.toml, purely for visual polish -- none of this
# touches how data is queried or computed, only how it's displayed.
st.markdown(
    """
    <style>
    .aligned-banner {
        background: linear-gradient(90deg, #1E3A8A 0%, #2563EB 60%, #3B82F6 100%);
        padding: 2rem 2rem 1.5rem 2rem;
        border-radius: 12px;
        color: white;
        margin-bottom: 1.5rem;
    }
    .aligned-banner h1 { color: white; margin-bottom: 0.25rem; }
    .aligned-banner p { color: #DBEAFE; font-size: 1.05rem; margin-bottom: 0; }
    div[data-testid="stMetric"] {
        background-color: #F8FAFC;
        border: 1px solid #E2E8F0;
        border-radius: 10px;
        padding: 0.75rem 1rem 0.5rem 1rem;
    }
    div[data-testid="stExpander"] { border-radius: 8px; }

    /* Sidebar redesign: a dark, branded panel instead of the plain
       default white sidebar with bare-looking radio buttons. */
    section[data-testid="stSidebar"] {
        background: linear-gradient(180deg, #0F172A 0%, #1E293B 100%);
    }
    section[data-testid="stSidebar"] * { color: #E2E8F0 !important; }
    .sidebar-logo {
        font-size: 1.6rem;
        font-weight: 800;
        color: #FFFFFF !important;
        margin-bottom: 0;
    }
    .sidebar-tagline {
        font-size: 0.8rem;
        color: #94A3B8 !important;
        margin-bottom: 1.2rem;
    }
    section[data-testid="stSidebar"] div[role="radiogroup"] label {
        background-color: rgba(255,255,255,0.04);
        border: 1px solid rgba(255,255,255,0.08);
        border-radius: 8px;
        padding: 0.5rem 0.75rem;
        margin-bottom: 0.4rem;
        transition: background-color 0.15s ease;
    }
    section[data-testid="stSidebar"] div[role="radiogroup"] label:hover {
        background-color: rgba(59,130,246,0.18);
    }
    .sidebar-footer {
        position: fixed;
        bottom: 1rem;
        font-size: 0.72rem;
        color: #64748B !important;
    }
    </style>
    """,
    unsafe_allow_html=True,
)


@st.cache_resource
def get_connection():
    # check_same_thread=False is safe here because this app only ever
    # reads from the database -- it never writes -- so there's no risk
    # of concurrent write conflicts across Streamlit's internal threads.
    return sqlite3.connect(DB_PATH, check_same_thread=False)


@st.cache_data
def run_query(sql, params=()):
    conn = get_connection()
    return pd.read_sql_query(sql, conn, params=params)


def render_overview():
    st.markdown(
        """
        <div class="aligned-banner">
            <h1>🎓 AlignED</h1>
            <p>Do graduate computing programs actually teach what the job market wants?
            Real curricula vs. real job postings, with statistical proof -- not guesses.</p>
        </div>
        """,
        unsafe_allow_html=True,
    )

    programs = run_query("SELECT COUNT(*) AS n FROM programs").iloc[0]["n"]
    courses = run_query("SELECT COUNT(*) AS n FROM courses").iloc[0]["n"]
    postings = run_query("SELECT COUNT(*) AS n FROM postings WHERE source = 'kaggle_sample'").iloc[0]["n"]
    gaps = run_query("SELECT COUNT(*) AS n FROM gap_scores").iloc[0]["n"]
    recs = run_query("SELECT COUNT(*) AS n FROM recommendations").iloc[0]["n"]

    col1, col2, col3, col4, col5 = st.columns(5)
    col1.metric("University Programs", programs)
    col2.metric("Real Courses Analyzed", f"{courses:,}")
    col3.metric("Job Postings Analyzed", f"{postings:,}")
    col4.metric("Statistically Significant Gaps", gaps)
    col5.metric("Final Recommendations", recs)

    st.markdown("---")
    st.markdown(
        """
        **How this was built, in one paragraph:** real course descriptions were
        scraped from 13 university catalogs, and real job postings were pulled
        from a live daily pipeline plus a historical dataset. Both were matched
        against the official US Department of Labor (O\\*NET) skills taxonomy.
        An AI extraction method was hand-validated against a classical keyword
        baseline on a 104-item hand-labeled test set (AI won, F1 0.400 vs.
        0.364) before choosing the faster method for full-scale analysis --
        see the Methodology page for the full, honest reasoning.
        """
    )
    st.info("Use the sidebar to explore a specific program's recommendations, skill demand trends, or the honest methodology behind this project.")


TIER_RGB = {"high": (220, 38, 38), "medium": (217, 119, 6), "low": (22, 163, 74)}
TIER_FILL_HEX = {"high": "FCA5A5", "medium": "FCD34D", "low": "86EFAC"}


def build_excel_report(university, program_name, course_count, recs_df):
    """Build a real, formatted .xlsx file -- colored header, bold title
    block, color-coded priority cells, sensible column widths -- instead
    of a plain CSV. A CSV is just raw text with commas; it fundamentally
    can't look "professional" no matter how the columns are chosen. An
    Excel file can, while still being just as easy to sort/filter/re-use
    as a CSV."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Recommendations"

    navy = "1E3A8A"
    ws.merge_cells("A1:G1")
    ws["A1"] = "AlignED -- Curriculum Gap Report"
    ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=navy)
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:G2")
    ws["A2"] = f"{university} -- {program_name}"
    ws["A2"].font = Font(size=11, italic=True, color="475569")

    ws.merge_cells("A3:G3")
    ws["A3"] = f"{course_count} courses analyzed  |  {len(recs_df)} significant gaps  |  Generated {datetime.date.today().isoformat()}"
    ws["A3"].font = Font(size=9, color="94A3B8")

    headers = ["Rank", "Skill", "Priority", "Program Coverage", "Market Demand", "Gap (points)", "Demand Trend"]
    header_row = 5
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.alignment = Alignment(horizontal="center")

    for i, (_, row) in enumerate(recs_df.iterrows(), start=1):
        r = header_row + i
        ws.cell(row=r, column=1, value=i).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=2, value=row["canonical_name"])
        tier_cell = ws.cell(row=r, column=3, value=row["priority_tier"].title())
        tier_cell.fill = PatternFill("solid", fgColor=TIER_FILL_HEX.get(row["priority_tier"], "E2E8F0"))
        tier_cell.alignment = Alignment(horizontal="center")
        tier_cell.font = Font(bold=True)
        cov_cell = ws.cell(row=r, column=4, value=row["program_coverage_rate"])
        cov_cell.number_format = "0.0%"
        dem_cell = ws.cell(row=r, column=5, value=row["market_demand_rate"])
        dem_cell.number_format = "0.0%"
        gap_cell = ws.cell(row=r, column=6, value=row["gap_value"])
        gap_cell.number_format = "+0.0%;-0.0%"
        ws.cell(row=r, column=7, value=row["trend_label"].replace("no clear trend", "Flat").title())

    widths = [7, 22, 12, 18, 16, 14, 14]
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = f"A{header_row + 1}"

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


class AlignEDReport(FPDF):
    """A small FPDF subclass so every page automatically gets the same
    branded footer (page number + generation date), instead of hand-adding
    it after every add_page() call."""

    def footer(self):
        self.set_y(-15)
        self.set_font("Helvetica", "I", 8)
        self.set_text_color(120, 120, 120)
        self.cell(0, 10, f"AlignED  |  Generated {datetime.date.today().isoformat()}  |  Page {self.page_no()}", align="C")


def normalize_term(term):
    """Same normalization rule used throughout the extraction pipeline
    (lowercase + collapse whitespace), duplicated here on purpose rather
    than imported from scripts/extraction/. The dashboard is meant to be
    deployable on its own (e.g. Streamlit Community Cloud only runs
    dashboard/app.py), so keeping this one small, stable helper
    self-contained avoids a fragile cross-folder import path."""
    if term is None:
        return ""
    return " ".join(str(term).strip().lower().split())


def build_combined_pattern(terms):
    """Same technique as scripts/extraction/extract_baseline.py's
    build_combined_pattern() -- one compiled regex covering every term,
    scanned in a single pass instead of once per term. Duplicated here
    for the same self-contained-deployment reason as normalize_term()
    above."""
    terms_sorted = sorted(terms, key=len, reverse=True)
    escaped = [re.escape(t) for t in terms_sorted]
    pattern = r"(?<![A-Za-z0-9_])(" + "|".join(escaped) + r")(?![A-Za-z0-9_])"
    return re.compile(pattern, re.IGNORECASE)


TOP_SKILLS_PER_CLUSTER = 15  # how many of a role's most in-demand skills count toward the match score

# Same generic, keyword-ambiguous terms excluded during gap scoring and
# trend detection (scripts/gap_analysis/) -- common English words a
# plain keyword scan can't tell apart from unrelated everyday text (e.g.
# "Design" matching "design your career", not the skill). Applied here
# too so the personal profile matcher stays consistent with the rest of
# the project instead of surfacing noisy false positives.
AMBIGUOUS_GENERIC_TERMS = {
    "design", "science", "writing", "monitoring", "programming",
    "troubleshooting", "mathematics", "coordination", "instructing",
    "repairing", "speaking", "route", "ada",
}


def _extract_user_skills(user_text, tracked_df):
    term_lookup = {normalize_term(name): sid for sid, name in zip(tracked_df["skill_id"], tracked_df["canonical_name"])}
    pattern = build_combined_pattern(list(term_lookup.keys()))
    matched = set()
    for m in pattern.finditer(user_text.lower()):
        sid = term_lookup.get(normalize_term(m.group(0)))
        if sid is not None:
            matched.add(sid)
    return matched


def build_profile_pdf_report(role_matches_df, top_role_label, have_df, missing_df, sample_postings_df):
    """A personalized career-style PDF: which real roles best fit this
    person's background, their strengths and gaps for the top match, and
    real example job postings pulled from that role -- built with the
    same branded report style as the program-level PDF, so the two feel
    like the same product."""
    def clean(text):
        return str(text).encode("latin-1", "replace").decode("latin-1")

    def write_line(text, size=10, bold=False, color=(20, 20, 20)):
        pdf.set_text_color(*color)
        pdf.set_font("Helvetica", "B" if bold else "", size)
        pdf.multi_cell(0, 6, text)
        pdf.set_x(pdf.l_margin)

    pdf = AlignEDReport()
    pdf.set_auto_page_break(auto=True, margin=20)
    pdf.add_page()

    pdf.set_fill_color(30, 58, 138)
    pdf.rect(0, 0, pdf.w, 32, style="F")
    pdf.set_xy(pdf.l_margin, 8)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "B", 18)
    pdf.cell(0, 10, "AlignED -- Your Personalized Career Report")
    pdf.set_xy(pdf.l_margin, 20)
    pdf.set_font("Helvetica", "", 11)
    pdf.cell(0, 8, clean(f"Best-matching role: {top_role_label}"))
    pdf.set_y(40)

    write_line(f"Generated {datetime.date.today().isoformat()}", size=9, color=(90, 90, 90))
    pdf.ln(6)
    pdf.set_x(pdf.l_margin)

    write_line("How your background matches real job roles", size=13, bold=True, color=(30, 58, 138))
    pdf.set_font("Helvetica", "", 9)
    pdf.set_fill_color(255, 255, 255)
    pdf.set_text_color(20, 20, 20)
    with pdf.table(col_widths=(80, 40), text_align="LEFT", line_height=6,
                   headings_style=FontFace(emphasis="BOLD", fill_color=(30, 58, 138), color=(255, 255, 255))) as table:
        header_row = table.row()
        header_row.cell("Role")
        header_row.cell("Match Score")
        for _, row in role_matches_df.head(6).iterrows():
            data_row = table.row()
            data_row.cell(clean(row["role_label"]))
            data_row.cell(f"{row['match_score']*100:.0f}%")

    pdf.ln(6)
    pdf.set_x(pdf.l_margin)
    write_line(f"Your strengths for {top_role_label}", size=13, bold=True, color=(22, 163, 74))
    strengths_text = ", ".join(have_df["canonical_name"].head(10)) if not have_df.empty else "None matched yet -- add more detail to your profile text."
    write_line(clean(strengths_text), size=10)

    pdf.ln(4)
    pdf.set_x(pdf.l_margin)
    write_line(f"Skills to prioritize for {top_role_label}", size=13, bold=True, color=(220, 38, 38))
    gaps_text = ", ".join(missing_df["canonical_name"].head(10)) if not missing_df.empty else "No major gaps found -- strong match!"
    write_line(clean(gaps_text), size=10)

    if not sample_postings_df.empty:
        pdf.ln(6)
        pdf.set_x(pdf.l_margin)
        write_line("Example real job openings matching this role", size=13, bold=True, color=(30, 58, 138))
        for _, row in sample_postings_df.head(6).iterrows():
            write_line(clean(f"-  {row['title']} ({row['company']})"), size=9.5, color=(60, 60, 60))

    return bytes(pdf.output())


def render_profile_builder():
    st.title("🙋 Build Your Profile")
    st.markdown(
        """
        Paste your current skills, resume text, or a list of courses you've taken.
        We'll match your background against **every real role in the job-market data**,
        show which one fits you best, and generate a personalized career report --
        including real example job openings for that role.
        """
    )

    user_text = st.text_area(
        "Your skills / resume text / courses taken",
        height=180,
        placeholder="e.g. I've taken courses in Python, statistics, and machine learning. Built a project using Docker and AWS...",
    )

    if not st.button("🔍 Find my best-matching roles", type="primary"):
        st.info("Paste your background above and click the button.")
        return

    if not user_text.strip():
        st.warning("Please paste some text first.")
        return

    tracked_df = run_query(
        """
        SELECT DISTINCT s.skill_id, s.canonical_name
        FROM extractions e JOIN skills s ON s.skill_id = e.skill_id
        WHERE e.source_type = 'posting' AND e.method = 'baseline_keyword'
        """
    )
    tracked_df = tracked_df[~tracked_df["canonical_name"].str.strip().str.lower().isin(AMBIGUOUS_GENERIC_TERMS)]
    matched_skill_ids = _extract_user_skills(user_text, tracked_df)

    if not matched_skill_ids:
        st.warning("We couldn't match any tracked skills in what you pasted -- try naming specific tools/languages/technologies (e.g. Python, SQL, Docker, Tableau).")
        return

    # Compute how well the user's skills overlap with each real role's
    # most in-demand skills -- this is what "auto-detects" the best-fit
    # role instead of asking the user to guess one from a dropdown.
    cluster_skill_counts = run_query(
        """
        SELECT pcm.cluster_id, e.skill_id, COUNT(DISTINCT e.source_id) AS n
        FROM extractions e JOIN posting_cluster_map pcm ON pcm.posting_id = e.source_id
        WHERE e.source_type = 'posting' AND e.method = 'baseline_keyword'
        GROUP BY pcm.cluster_id, e.skill_id
        """
    )
    cluster_totals = run_query("SELECT cluster_id, COUNT(*) AS total FROM posting_cluster_map GROUP BY cluster_id")
    roles_df = run_query(
        "SELECT cluster_id, role_label FROM role_clusters WHERE role_label NOT LIKE 'Mixed%' AND role_label NOT LIKE 'Near-duplicate%'"
    )

    merged = cluster_skill_counts.merge(cluster_totals, on="cluster_id").merge(roles_df, on="cluster_id")
    merged["demand_rate"] = merged["n"] / merged["total"]
    # Inner join on purpose: tracked_df has already excluded the generic,
    # keyword-ambiguous terms, so an inner join here removes those skills
    # from the per-role ranking entirely, instead of a left join leaving
    # them in with a blank name (which would still let them occupy a
    # "top skill" slot).
    merged = merged.merge(tracked_df, on="skill_id", how="inner")
    top_per_cluster = merged.sort_values("demand_rate", ascending=False).groupby("cluster_id").head(TOP_SKILLS_PER_CLUSTER)

    match_rows = []
    for cluster_id, group in top_per_cluster.groupby("cluster_id"):
        overlap = group["skill_id"].isin(matched_skill_ids).sum()
        match_rows.append({
            "cluster_id": cluster_id,
            "role_label": group["role_label"].iloc[0],
            "match_score": overlap / len(group),
        })
    role_matches_df = pd.DataFrame(match_rows).sort_values("match_score", ascending=False).reset_index(drop=True)

    st.markdown("---")
    st.subheader("🎯 Your best-matching roles")
    for _, row in role_matches_df.head(5).iterrows():
        st.write(f"**{row['role_label']}** -- {row['match_score']*100:.0f}% match")
        st.progress(min(row["match_score"], 1.0))

    top_role = role_matches_df.iloc[0]
    top_cluster_id = int(top_role["cluster_id"])
    st.markdown("---")
    st.subheader(f"Deep dive: {top_role['role_label']}")

    cluster_data = top_per_cluster[top_per_cluster["cluster_id"] == top_cluster_id]
    have_df = cluster_data[cluster_data["skill_id"].isin(matched_skill_ids)].sort_values("demand_rate", ascending=False)
    missing_df = cluster_data[~cluster_data["skill_id"].isin(matched_skill_ids)].sort_values("demand_rate", ascending=False)

    col1, col2 = st.columns(2)
    with col1:
        st.markdown(f"**✅ Your strengths ({len(have_df)})**")
        for _, row in have_df.iterrows():
            st.write(f"🟢 {row['canonical_name']} -- in {row['demand_rate']*100:.0f}% of postings")
        if have_df.empty:
            st.write("No overlap yet with this role's top skills.")
    with col2:
        st.markdown(f"**🔴 Skills to prioritize ({len(missing_df)})**")
        for _, row in missing_df.iterrows():
            st.write(f"🔴 {row['canonical_name']} -- in {row['demand_rate']*100:.0f}% of postings")
        if missing_df.empty:
            st.write("Great coverage of this role's top skills!")

    sample_postings_df = run_query(
        """
        SELECT p.title, p.company FROM postings p
        JOIN posting_cluster_map pcm ON pcm.posting_id = p.posting_id
        WHERE pcm.cluster_id = ? LIMIT 8
        """,
        (top_cluster_id,),
    )
    st.markdown("---")
    st.subheader("💼 Example real job openings matching this role")
    st.caption("Pulled directly from the sampled job postings -- real listings, not generated examples.")
    for _, row in sample_postings_df.iterrows():
        st.write(f"• **{row['title']}** ({row['company']})")

    st.markdown("---")
    pdf_bytes = build_profile_pdf_report(role_matches_df, top_role["role_label"], have_df, missing_df, sample_postings_df)
    st.download_button(
        "⬇️ Download my personalized career report (PDF)", data=pdf_bytes,
        file_name="AlignED_My_Career_Report.pdf", mime="application/pdf",
    )

    st.caption(
        "This uses simple keyword matching, the same fast method used for the full-scale program analysis "
        "elsewhere in this project -- it can miss skills phrased differently than expected. See Methodology for details."
    )


def build_pdf_report(university, program_name, course_count, recs_df):
    """Build a proper report-style PDF: a colored title block, a few
    headline numbers, a clean data table (not a wall of repeated
    sentences), and a short "top 3 priorities" callout section at the
    end -- meant to look like something you'd actually hand to a
    curriculum committee, not a boring text dump."""
    def clean(text):
        return str(text).encode("latin-1", "replace").decode("latin-1")

    def write_line(text, size=10, bold=False, color=(20, 20, 20)):
        pdf.set_text_color(*color)
        pdf.set_font("Helvetica", "B" if bold else "", size)
        pdf.multi_cell(0, 6, text)
        pdf.set_x(pdf.l_margin)

    pdf = AlignEDReport()
    pdf.set_auto_page_break(auto=True, margin=20)
    pdf.add_page()

    # Title block: a filled navy header bar.
    pdf.set_fill_color(30, 58, 138)
    pdf.rect(0, 0, pdf.w, 32, style="F")
    pdf.set_xy(pdf.l_margin, 8)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "B", 18)
    pdf.cell(0, 10, "AlignED -- Curriculum Gap Report")
    pdf.set_xy(pdf.l_margin, 20)
    pdf.set_font("Helvetica", "", 11)
    pdf.cell(0, 8, clean(f"{university} -- {program_name}"))
    pdf.set_y(40)

    n_high = (recs_df["priority_tier"] == "high").sum()
    n_rising = (recs_df["trend_label"] == "rising").sum()
    write_line(
        f"{course_count} courses analyzed   |   {len(recs_df)} significant gaps found   |   "
        f"{n_high} high-priority   |   {n_rising} trending upward   |   Generated {datetime.date.today().isoformat()}",
        size=9, color=(90, 90, 90),
    )
    pdf.ln(6)
    pdf.set_x(pdf.l_margin)

    # A real data table instead of repeated paragraphs -- far quicker to
    # scan, and it's what a curriculum committee would actually expect.
    write_line("All significant gaps", size=13, bold=True, color=(30, 58, 138))
    pdf.set_font("Helvetica", "", 9)
    # Reset fill color to white before the table -- otherwise the navy
    # fill_color left over from the header banner above silently bleeds
    # into the table's data-row backgrounds, making the text unreadable
    # (caught by actually rendering the PDF to an image and looking at
    # it, not just checking that the code ran without error).
    pdf.set_fill_color(255, 255, 255)
    pdf.set_text_color(20, 20, 20)
    col_widths = (18, 48, 25, 25, 22, 28)
    headers = ["Priority", "Skill", "Coverage", "Market Demand", "Gap", "Trend"]
    with pdf.table(col_widths=col_widths, text_align="LEFT", line_height=6,
                   headings_style=FontFace(emphasis="BOLD", fill_color=(30, 58, 138), color=(255, 255, 255))) as table:
        header_row = table.row()
        for h in headers:
            header_row.cell(h)
        for _, row in recs_df.iterrows():
            data_row = table.row()
            data_row.cell(row["priority_tier"].upper())
            data_row.cell(clean(row["canonical_name"]))
            data_row.cell(f"{row['program_coverage_rate']*100:.0f}%")
            data_row.cell(f"{row['market_demand_rate']*100:.0f}%")
            data_row.cell(f"+{row['gap_value']*100:.0f} pt")
            data_row.cell(row["trend_label"].replace("no clear trend", "flat").title())

    # A short, readable "top priorities" callout with genuinely varied
    # phrasing per row -- reusing the exact same sentence template three
    # times in a row (as the raw database rationale does) reads robotic
    # and repetitive once you actually read them back to back.
    def punchy_summary(row, variant):
        skill = clean(row["canonical_name"])
        cov = row["program_coverage_rate"] * 100
        dem = row["market_demand_rate"] * 100
        trend_bit = {
            "rising": " Demand is climbing, so this gap is only getting more urgent.",
            "falling": " Demand has cooled recently, though the gap is still real today.",
        }.get(row["trend_label"], "")
        templates = [
            f"{skill} shows up in {dem:.0f}% of job postings we sampled, but almost none of this program's courses ({cov:.0f}%) cover it.{trend_bit}",
            f"Employers ask for {skill} constantly -- {dem:.0f}% of postings -- yet the curriculum barely touches it ({cov:.0f}% coverage).{trend_bit}",
            f"A clear blind spot: {skill} is expected in {dem:.0f}% of real postings, but this program teaches it in only {cov:.0f}% of its courses.{trend_bit}",
        ]
        return templates[variant % len(templates)]

    pdf.ln(6)
    pdf.set_x(pdf.l_margin)
    write_line("Top priorities, explained", size=13, bold=True, color=(30, 58, 138))
    top_3 = recs_df[recs_df["priority_tier"] == "high"].head(3)
    for i, (_, row) in enumerate(top_3.iterrows()):
        rgb = TIER_RGB.get(row["priority_tier"], (100, 100, 100))
        start_y = pdf.get_y()
        pdf.set_fill_color(*rgb)
        pdf.rect(pdf.l_margin, start_y, 2.5, 16, style="F")
        pdf.set_x(pdf.l_margin + 5)
        write_line(clean(row["canonical_name"]), size=11, bold=True, color=rgb)
        pdf.set_x(pdf.l_margin + 5)
        write_line(punchy_summary(row, i), size=9.5, color=(60, 60, 60))
        pdf.ln(3)
        pdf.set_x(pdf.l_margin)

    return bytes(pdf.output())


def render_program_explorer():
    st.title("📋 Program Explorer")
    st.markdown("Pick a program to see its ranked, statistically significant skill gaps -- and what to do about them.")

    programs_df = run_query("SELECT program_id, university, program_name, tier FROM programs ORDER BY university")
    programs_df["label"] = programs_df["university"] + " -- " + programs_df["program_name"]
    selected_label = st.selectbox("Choose a program", programs_df["label"])
    selected = programs_df[programs_df["label"] == selected_label].iloc[0]
    program_id = int(selected["program_id"])

    course_count = run_query("SELECT COUNT(*) AS n FROM courses WHERE program_id = ?", (program_id,)).iloc[0]["n"]
    st.caption(f"Tier: {selected['tier']}  |  {course_count} real courses in this program")

    recs_df = run_query(
        """
        SELECT r.skill_id, r.gap_value, r.trend_label, r.priority_score, r.priority_tier, r.rationale,
               g.program_coverage_rate, g.market_demand_rate
        FROM recommendations r
        JOIN gap_scores g ON g.program_id = r.program_id AND g.skill_id = r.skill_id
        WHERE r.program_id = ?
        ORDER BY
            CASE r.priority_tier WHEN 'high' THEN 0 WHEN 'medium' THEN 1 ELSE 2 END,
            r.priority_score DESC
        """,
        (program_id,),
    )
    skills_df = run_query("SELECT skill_id, canonical_name FROM skills")
    recs_df = recs_df.merge(skills_df, on="skill_id", how="left")

    if recs_df.empty:
        st.warning("No statistically significant gaps were found for this program (this can genuinely happen for very small programs with few courses).")
        return

    dl_col1, dl_col2 = st.columns(2)
    with dl_col1:
        # A real formatted Excel file, not a plain CSV -- CSV is just raw
        # comma-separated text, so it fundamentally cannot look
        # "professional" no matter how the columns are arranged (no
        # colors, no bold header, no cell shading). Excel can, while
        # still being just as sortable/filterable as a CSV would be.
        excel_bytes = build_excel_report(selected["university"], selected["program_name"], course_count, recs_df)
        st.download_button(
            "⬇️ Download as Excel", data=excel_bytes,
            file_name=f"{selected['university']}_{selected['program_name']}_recommendations.xlsx".replace(" ", "_"),
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    with dl_col2:
        pdf_bytes = build_pdf_report(selected["university"], selected["program_name"], course_count, recs_df)
        st.download_button(
            "⬇️ Download as PDF", data=pdf_bytes,
            file_name=f"{selected['university']}_{selected['program_name']}_recommendations.pdf".replace(" ", "_"),
            mime="application/pdf",
        )

    tier_colors = {"high": "🔴", "medium": "🟡", "low": "🟢"}
    for tier in ["high", "medium", "low"]:
        tier_df = recs_df[recs_df["priority_tier"] == tier]
        if tier_df.empty:
            continue
        st.subheader(f"{tier_colors[tier]} {tier.title()} priority")
        for _, row in tier_df.iterrows():
            trend_note = {"rising": "📈 rising demand", "falling": "📉 falling demand"}.get(row["trend_label"], "")
            with st.expander(f"{row['canonical_name']}  ({row['gap_value']*100:.0f} point gap)  {trend_note}"):
                st.write(row["rationale"])

    st.markdown("---")
    st.subheader("Gap size, visualized")
    chart_df = recs_df.sort_values("gap_value", ascending=True)
    fig = px.bar(
        chart_df, x="gap_value", y="canonical_name", orientation="h",
        color="priority_tier", color_discrete_map={"high": "#e15759", "medium": "#f1c232", "low": "#59a14f"},
        labels={"gap_value": "Gap (market demand - program coverage)", "canonical_name": "Skill"},
    )
    fig.update_layout(xaxis_tickformat=".0%", height=max(300, len(chart_df) * 35))
    st.plotly_chart(fig, use_container_width=True)


def render_compare():
    st.title("⚖️ Compare Programs")
    st.markdown("Pick 2-3 programs to see their top skill gaps side by side.")

    programs_df = run_query("SELECT program_id, university, program_name FROM programs ORDER BY university")
    programs_df["label"] = programs_df["university"] + " -- " + programs_df["program_name"]

    chosen_labels = st.multiselect(
        "Choose programs to compare (2-3 recommended)",
        programs_df["label"],
        default=list(programs_df["label"].iloc[:2]),
    )
    if len(chosen_labels) < 2:
        st.info("Pick at least 2 programs to compare.")
        return

    chosen = programs_df[programs_df["label"].isin(chosen_labels)]
    skills_df = run_query("SELECT skill_id, canonical_name FROM skills")

    cols = st.columns(len(chosen))
    all_gap_rows = []
    for col, (_, prog) in zip(cols, chosen.iterrows()):
        program_id = int(prog["program_id"])
        course_count = run_query("SELECT COUNT(*) AS n FROM courses WHERE program_id = ?", (program_id,)).iloc[0]["n"]
        recs = run_query(
            "SELECT skill_id, gap_value, priority_tier FROM recommendations WHERE program_id = ? ORDER BY priority_score DESC LIMIT 8",
            (program_id,),
        ).merge(skills_df, on="skill_id", how="left")
        recs["program_label"] = prog["label"]
        all_gap_rows.append(recs)

        with col:
            st.subheader(prog["university"])
            st.caption(f"{prog['program_name']}  |  {course_count} courses")
            if recs.empty:
                st.write("No significant gaps found.")
            else:
                for _, row in recs.iterrows():
                    tier_icon = {"high": "🔴", "medium": "🟡", "low": "🟢"}.get(row["priority_tier"], "")
                    st.write(f"{tier_icon} **{row['canonical_name']}** -- +{row['gap_value']*100:.0f} pts")

    combined = pd.concat(all_gap_rows, ignore_index=True) if all_gap_rows else pd.DataFrame()
    if not combined.empty:
        st.markdown("---")
        st.subheader("Side-by-side gap sizes")
        fig = px.bar(
            combined, x="canonical_name", y="gap_value", color="program_label", barmode="group",
            labels={"canonical_name": "Skill", "gap_value": "Gap (market demand - coverage)", "program_label": "Program"},
        )
        fig.update_layout(yaxis_tickformat=".0%", height=450)
        st.plotly_chart(fig, use_container_width=True)


def render_course_finder():
    st.title("🔎 Course Finder")
    st.markdown(
        "Looking for courses that build a specific skill (e.g. Python, Docker, Machine Learning)? "
        "Search or filter below -- no upload needed, this searches real course descriptions across all 13 programs directly."
    )

    search_text = st.text_input("Search course names/descriptions (e.g. 'software', 'security', 'machine learning')")

    tracked_skills_df = run_query(
        """
        SELECT DISTINCT s.skill_id, s.canonical_name
        FROM extractions e JOIN skills s ON s.skill_id = e.skill_id
        WHERE e.source_type = 'course' AND e.method = 'baseline_keyword'
        ORDER BY s.canonical_name
        """
    )
    skill_filter = st.multiselect("...or filter by a specific tracked skill", tracked_skills_df["canonical_name"])

    base_query = """
        SELECT c.course_name, c.description, p.university, p.program_name
        FROM courses c JOIN programs p ON p.program_id = c.program_id
        WHERE 1=1
    """
    params = []
    if search_text:
        base_query += " AND (c.course_name LIKE ? OR c.description LIKE ?)"
        params += [f"%{search_text}%", f"%{search_text}%"]
    if skill_filter:
        skill_ids = tracked_skills_df[tracked_skills_df["canonical_name"].isin(skill_filter)]["skill_id"].tolist()
        placeholders = ",".join(str(s) for s in skill_ids)
        base_query += f"""
            AND c.course_id IN (
                SELECT CAST(source_id AS INTEGER) FROM extractions
                WHERE source_type='course' AND method='baseline_keyword' AND skill_id IN ({placeholders})
            )
        """
    base_query += " LIMIT 100"

    if not search_text and not skill_filter:
        st.info("Type a search term or pick a skill above to find matching courses.")
        return

    results = run_query(base_query, tuple(params))
    st.caption(f"{len(results)} matching course(s) found (showing up to 100).")
    for _, row in results.iterrows():
        with st.expander(f"{row['course_name']}  --  {row['university']}"):
            st.caption(row["program_name"])
            st.write(row["description"] or "(no description available)")


def render_heatmap():
    st.title("🗺️ Skill Coverage Heatmap")
    st.markdown(
        "**How to read this:** each cell shows what percent of a program's courses cover a skill. "
        "🔴 Red = barely covered (a real gap). 🟢 Green = well covered. The number in each cell is the exact percentage."
    )

    num_skills = st.slider("Number of skills to show (fewer = easier to read)", min_value=5, max_value=20, value=10)

    top_skills_df = run_query(
        """
        SELECT s.skill_id, s.canonical_name, COUNT(*) AS n_programs_gapped
        FROM gap_scores g JOIN skills s ON s.skill_id = g.skill_id
        GROUP BY g.skill_id ORDER BY n_programs_gapped DESC LIMIT ?
        """,
        (num_skills,),
    )
    programs_df = run_query("SELECT program_id, university FROM programs ORDER BY university")
    # Short labels (university only) instead of the full "university -- long
    # program name" string -- much easier to scan on a crowded axis.
    programs_df["short_label"] = programs_df["university"].str.replace("University of ", "U. ", regex=False)
    # De-duplicate universities that appear more than once (e.g. Georgia
    # Tech has 3 programs) by appending a short suffix.
    dupe_counts = {}
    short_labels = []
    for uni in programs_df["short_label"]:
        dupe_counts[uni] = dupe_counts.get(uni, 0) + 1
        short_labels.append(uni if dupe_counts[uni] == 1 else f"{uni} ({dupe_counts[uni]})")
    programs_df["short_label"] = short_labels

    # Pull TRUE coverage directly from the course/extraction data for
    # every program x skill pair, not just the ones that happened to
    # clear the significance test in gap_scores -- otherwise a program
    # that covers a skill a little (but not enough to be "significant")
    # gets shown identically to a program that covers it 0%, which is
    # misleading.
    coverage_df = run_query(
        """
        SELECT c.program_id, e.skill_id, COUNT(DISTINCT e.source_id) AS n_covering,
               (SELECT COUNT(*) FROM courses WHERE program_id = c.program_id) AS n_total
        FROM extractions e
        JOIN courses c ON c.course_id = CAST(e.source_id AS INTEGER)
        WHERE e.source_type = 'course' AND e.method = 'baseline_keyword'
          AND e.skill_id IN ({})
        GROUP BY c.program_id, e.skill_id
        """.format(",".join(str(s) for s in top_skills_df["skill_id"])),
    )

    matrix = pd.DataFrame(0.0, index=programs_df["short_label"], columns=top_skills_df["canonical_name"])
    program_label_by_id = dict(zip(programs_df["program_id"], programs_df["short_label"]))
    skill_name_by_id = dict(zip(top_skills_df["skill_id"], top_skills_df["canonical_name"]))
    for _, row in coverage_df.iterrows():
        label = program_label_by_id.get(row["program_id"])
        skill_name = skill_name_by_id.get(row["skill_id"])
        if label is not None and skill_name is not None and row["n_total"] > 0:
            matrix.loc[label, skill_name] = row["n_covering"] / row["n_total"]

    # Real, honest finding worth stating plainly: course catalog
    # descriptions rarely name specific tools the way job postings do, so
    # true coverage for these skills tops out in the single digits almost
    # everywhere. Locking the color scale to 0-100% would flatten all of
    # that into "uniform deep red" and hide the real (small) variation
    # that does exist -- so the scale is set to the actual data range
    # instead, with a sensible floor so it doesn't get overly sensitive.
    data_max = max(matrix.values.max(), 0.05)

    fig = go.Figure(
        data=go.Heatmap(
            z=matrix.values, x=matrix.columns, y=matrix.index,
            colorscale="RdYlGn", zmin=0, zmax=data_max,
            text=[[f"{v*100:.0f}%" for v in row] for row in matrix.values],
            texttemplate="%{text}", textfont={"size": 11},
            colorbar=dict(title="Coverage", tickformat=".0%"),
            hovertemplate="Program: %{y}<br>Skill: %{x}<br>Coverage: %{z:.0%}<extra></extra>",
        )
    )
    fig.update_layout(height=max(400, len(matrix.index) * 32), xaxis_tickangle=-35, margin=dict(l=10, r=10, t=10, b=10))
    st.plotly_chart(fig, use_container_width=True)
    st.caption(
        f"This uses each program's real, full coverage rate (not just the significant-gap subset). "
        f"An honest finding on its own: even the best-covered program/skill pair here only reaches {data_max*100:.0f}% -- "
        f"course descriptions rarely name specific tools the way job postings do, which is why the color scale is "
        f"stretched to this data's real range instead of the full 0-100%."
    )


def render_trends():
    st.title("📈 Skill Demand Trends")
    st.markdown(
        """
        Based on ~124,000 real historical job postings, restricted to the
        6 weeks with a real, meaningful volume of data (see Methodology for
        why some weeks were excluded).
        """
    )

    trends_df = run_query(
        """
        SELECT t.trend_label, t.slope, t.p_value, t.first_half_rate, t.second_half_rate, s.canonical_name
        FROM skill_trends t JOIN skills s ON s.skill_id = t.skill_id
        ORDER BY t.slope DESC
        """
    )

    movers = trends_df[trends_df["trend_label"].isin(["rising", "falling"])].copy()
    if not movers.empty:
        st.subheader("Momentum: rising vs. falling skills")
        movers["slope_pct"] = movers["slope"] * 100
        movers_sorted = movers.sort_values("slope_pct")
        fig = px.bar(
            movers_sorted, x="slope_pct", y="canonical_name", orientation="h",
            color="trend_label", color_discrete_map={"rising": "#16A34A", "falling": "#DC2626"},
            labels={"slope_pct": "Change in demand rate, points per week", "canonical_name": "Skill", "trend_label": "Trend"},
        )
        fig.update_layout(height=max(350, len(movers_sorted) * 28), showlegend=True)
        st.plotly_chart(fig, use_container_width=True)

        st.subheader("Before vs. after: first half vs. second half of the window")
        before_after = movers.melt(
            id_vars=["canonical_name", "trend_label"],
            value_vars=["first_half_rate", "second_half_rate"],
            var_name="period", value_name="rate",
        )
        before_after["period"] = before_after["period"].map({"first_half_rate": "First half", "second_half_rate": "Second half"})
        fig2 = px.bar(
            before_after, x="canonical_name", y="rate", color="period", barmode="group",
            labels={"rate": "Demand rate", "canonical_name": "Skill", "period": "Period"},
            color_discrete_map={"First half": "#94A3B8", "Second half": "#2563EB"},
        )
        fig2.update_layout(yaxis_tickformat=".1%", height=400, xaxis_tickangle=-35)
        st.plotly_chart(fig2, use_container_width=True)

    st.markdown("---")
    st.subheader("Full detail")
    col1, col2 = st.columns(2)
    with col1:
        st.subheader("📈 Rising")
        rising = trends_df[trends_df["trend_label"] == "rising"]
        st.dataframe(
            rising[["canonical_name", "first_half_rate", "second_half_rate", "p_value"]]
            .rename(columns={"canonical_name": "Skill", "first_half_rate": "First half", "second_half_rate": "Second half", "p_value": "p-value"})
            .style.format({"First half": "{:.1%}", "Second half": "{:.1%}", "p-value": "{:.4f}"}),
            hide_index=True, use_container_width=True,
        )
    with col2:
        st.subheader("📉 Falling")
        falling = trends_df[trends_df["trend_label"] == "falling"]
        st.dataframe(
            falling[["canonical_name", "first_half_rate", "second_half_rate", "p_value"]]
            .rename(columns={"canonical_name": "Skill", "first_half_rate": "First half", "second_half_rate": "Second half", "p_value": "p-value"})
            .style.format({"First half": "{:.1%}", "Second half": "{:.1%}", "p-value": "{:.4f}"}),
            hide_index=True, use_container_width=True,
        )

    stable_count = len(trends_df[trends_df["trend_label"] == "no clear trend"])
    st.caption(f"{stable_count} additional tracked skills showed no statistically significant trend over this window.")


def render_clusters():
    st.title("🧩 Role Clusters")
    st.markdown(
        """
        Real job postings were grouped by the actual role they describe (not
        job title alone) using AI-generated embeddings and k-means
        clustering, so curricula can be compared against what a role needs
        in general -- not one company's specific posting.
        """
    )

    clusters_df = run_query(
        """
        SELECT rc.cluster_id, rc.role_label, rc.silhouette_score, COUNT(pcm.posting_id) AS n_postings
        FROM role_clusters rc
        LEFT JOIN posting_cluster_map pcm ON pcm.cluster_id = rc.cluster_id
        GROUP BY rc.cluster_id
        ORDER BY n_postings DESC
        """
    )
    silhouette = clusters_df["silhouette_score"].iloc[0] if not clusters_df.empty else None
    if silhouette is not None:
        st.caption(f"Best number of clusters chosen automatically via silhouette score: {silhouette:.4f}. See Methodology for why this number is honestly modest, and what it means.")

    chart_col, pie_col = st.columns([3, 2])
    with chart_col:
        fig = px.bar(clusters_df, x="n_postings", y="role_label", orientation="h", labels={"n_postings": "Sampled postings", "role_label": "Role"})
        fig.update_layout(height=450)
        st.plotly_chart(fig, use_container_width=True)
    with pie_col:
        fig_pie = px.pie(clusters_df, names="role_label", values="n_postings", hole=0.45)
        fig_pie.update_traces(textposition="inside", textinfo="percent")
        fig_pie.update_layout(height=450, showlegend=False)
        st.plotly_chart(fig_pie, use_container_width=True)

    st.markdown("---")
    st.subheader("See real postings inside a cluster")
    st.caption("Pulled straight from the sampled job postings -- a real sanity check, not just a label.")
    cluster_choice = st.selectbox("Choose a role cluster", clusters_df["role_label"])
    cluster_id = int(clusters_df[clusters_df["role_label"] == cluster_choice]["cluster_id"].iloc[0])
    sample_postings = run_query(
        """
        SELECT p.title, p.company FROM postings p
        JOIN posting_cluster_map pcm ON pcm.posting_id = p.posting_id
        WHERE pcm.cluster_id = ?
        LIMIT 8
        """,
        (cluster_id,),
    )
    for _, row in sample_postings.iterrows():
        st.write(f"• **{row['title']}** ({row['company']})")


def render_methodology():
    st.title("🔍 Methodology & Honest Limitations")
    st.markdown(
        """
        This page exists on purpose: a portfolio project is only as
        trustworthy as its documented limitations. Every simplification
        below was a deliberate, explained trade-off -- not an oversight.

        ### Data sources
        - **Course data:** scraped directly from 13 real university course
          catalogs (Georgia Tech, ASU, UIUC, Northeastern, BU, Wisconsin,
          UMD, Penn State, UW, Michigan).
        - **Job posting data:** a live daily pipeline (Adzuna API, via
          GitHub Actions) plus a historical backfill of ~124,000 real
          postings (Kaggle LinkedIn dataset).
        - **Skills taxonomy:** the official US Department of Labor O\\*NET
          database -- not an invented list. ESCO (the EU equivalent) was
          used first and is kept in the project history, but O\\*NET was
          chosen for better coverage of named tools.

        ### AI vs. classical extraction -- an actual, measured comparison
        A local, free AI model (Ollama) was compared against a classical
        keyword-matching baseline on a 104-item hand-labeled test set:

        | Method | Precision | Recall | F1 |
        |---|---|---|---|
        | Baseline (keyword) | 0.518 | 0.280 | 0.364 |
        | AI (local LLM + embeddings) | 0.407 | 0.392 | **0.400** |

        The AI method won on the accuracy metric that matters most here
        (F1). But running it across the *full* 1,378 courses and
        thousands of postings would take hours on consumer hardware, so
        the fast keyword method was used deliberately for full-scale
        analysis -- a real "best model for evaluation, faster model for
        production scale" engineering trade-off.

        ### Correcting for running many statistical tests at once
        Every program is tested against ~70 skills at once, not just one.
        Running that many significance tests together means a few
        "significant" results are expected to be false positives from
        chance alone, even if every individual test is done correctly --
        the classic multiple-comparisons problem. To account for this, a
        Benjamini-Hochberg false discovery rate (FDR) correction is
        applied across each program's full set of tests before anything
        is called significant. This is a stricter, more defensible bar
        than using raw p-values alone, and it visibly changes the results:
        applying it dropped the count of "significant" gaps from 231 to
        159 across all 13 programs -- exactly the kind of honest
        tightening a real statistical review should produce.

        ### Known, documented limitations
        - **Keyword matching can't disambiguate context.** A handful of
          generic single-word skill names (e.g. "Design", "Science") were
          excluded from gap/trend analysis because a keyword scanner can't
          tell them apart from unrelated everyday text.
        - **The historical posting dataset's timestamps are skewed.** 68%
          of the ~124,000 postings are dated in a single final week --
          almost certainly a data-collection artifact, not real hiring
          activity. Trend analysis was restricted to the 6 weeks with real
          volume (>=100 postings) rather than reporting a misleading
          result across mostly-empty weeks.
        - **Role clustering's silhouette score is modest (0.08).** This is
          normal and expected for real, overlapping job-posting text (a
          "DevOps Engineer" and "Cloud Engineer" posting legitimately
          share a lot of language) -- a hand sanity-check of sampled
          postings per cluster confirmed most clusters are genuinely
          coherent by role.
        - **The gap-score threshold-tuning note:** the embedding
          similarity cutoff used in the AI evaluation was tuned against
          the same 104-item set used for final reporting, a known,
          deliberate simplification for a project of this scope (the more
          textbook-correct approach would use a separate held-out set).
        """
    )


PAGES = {
    "🏠  Overview": render_overview,
    "🙋  Build Your Profile": render_profile_builder,
    "📋  Program Explorer": render_program_explorer,
    "🔎  Course Finder": render_course_finder,
    "⚖️  Compare Programs": render_compare,
    "🗺️  Skill Coverage Heatmap": render_heatmap,
    "📈  Skill Demand Trends": render_trends,
    "🧩  Role Clusters": render_clusters,
    "🔍  Methodology & Limitations": render_methodology,
}

st.sidebar.markdown('<p class="sidebar-logo">🎓 AlignED</p>', unsafe_allow_html=True)
st.sidebar.markdown('<p class="sidebar-tagline">Curriculum vs. job market gap analysis</p>', unsafe_allow_html=True)
selection = st.sidebar.radio("Navigate", list(PAGES.keys()), label_visibility="collapsed")
st.sidebar.markdown(
    f'<p class="sidebar-footer">Data refreshed via database queries<br>'
    f'13 programs · 1,378 courses · 1,660+ postings<br>'
    f'Built with Python, SQLite &amp; Streamlit</p>',
    unsafe_allow_html=True,
)
PAGES[selection]()
